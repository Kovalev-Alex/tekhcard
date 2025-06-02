import openpyxl as op
from datetime import datetime
from pprint import pprint

def time(func):
    """
    Декоратор 'Время выполнения функции.'
    """
    def wrapper(*ar, **kw):
        start = datetime.now()
        result = func(*ar, **kw)
        print(f'Время выполнения: {datetime.now()-start}')
        return result
    return wrapper

def round_val(value:int | float) -> int | float:
    """
    Округление вверх. \n
    Для значений около +-0.0001: учитываются 4 знака после запятой.\n
    Для остальных: 15 знаков после запятой.
    """
    if 0 < value < 1:
        val = str(value)[:4]
        dif = str(value)[5:] 
        if int(dif) > 0:
            return round(float(val) + 0.01, 2)
        return float(val)
    elif value < 100:
        dot = str(value).find('.')
        if dot > 0:
            val = str(value)[:dot+2]
            dif = str(value)[dot+3:] 
            if int(dif.zfill(5)) > 0:
                return round(float(val) + 0.1, 1)
            return float(val)
        return int(value)
    elif value >= 100:
        dot = str(value).find('.')
        if dot > 0:
            val = str(value)[:dot]
            dif = str(value)[dot+1:] 
            if int(dif) > 0:
                return int(val) + 1
            return int(val)
        return int(value)

def get_kokil(D:int | float) -> list:
    """
    Принимает на вход диаметр_кокиля.\n
    \tВозвращает:\n
    Список кокилей указанного диаметра (кортежи).\n
    """
    
    file = '//ad/технологи/Перечень центробежных изложниц.xlsx'
    book = op.load_workbook(file)
    sheet = book['Изложницы без бурта']
    
    lst1 = []
    for row in sheet.iter_rows(min_row=3, min_col=2, max_col=8, max_row=1159, values_only=True):
        if row[0] == D:
            if row not in lst1:
                lst1.append(row)
    
    return(sorted(lst1))

def massa_vtulki(d1:float | int, d2:float | int, len:float | int, ro:float,) -> float | int:
    """
    Вычисляет массу втулки в кг.\r\n
    d1 - Наружный диаметр.
    d2 - Внутренний диаметр.
    len - Длина.
    ro - Плотность металла.
    """
    return (d1*d1-d2*d2)*3.1415926*len*ro/4000000

def massa_vtulki_s_burtom(d1:float | int, d2:float | int, d3:float | int, len1:float | int, len2:float, ro:float) -> float | int:
    """
    Вычисляет массу втулки с буртом в кг.\r\n
    d1 - Диаметр бурта.
    d2 - Диаметр бочки.
    d3 - Диаметр внутренний.
    len1 - Длина бурта.
    len2 - Длина общая.
    ro - Плотность металла.
    """
    return (d1*d1-d3*d3)*3.1415926*len1*ro/4000000 + (d2*d2-d3*d3)*3.1415926*(len2-len1)*ro/4000000

def massa_konusnaya(d1:float | int, d2:float | int, d3:float | int, d4:float | int, len:float | int, ro:float) -> float | int:
    """
    Расчет массы конусной втулки.\n
    d1 - Наружный больший диаметр конуса.
    d2 - Внутренний больший диаметр конуса.
    d3 - Наружный меньший диаметр конуса.
    d4 - Внутренний меньший диаметр конуса.
    len - Длина конуса.
    ro - Плотность металла.
    """
    return ((d1/2)*(d1/2)+(d1/2)*(d3/2)+(d3/2)*(d3/2))*3.1415926*len*ro/3000000-((d2/2)*(d2/2)+(d2/2)*(d4/2)+(d4/2)*(d4/2))*3.1415926*len*ro/3000000

def massa_planki(a:float | int, b:float | int, c:float | int, ro:float) -> float | int:
    """
    Вычисляет массу планки в кг.\n
    a - Толщина.
    b - Ширина.
    c - Длина.
    ro - Плотность металла.

    """
    return a*b*c*ro/1000000

def massa_bolvana_cb(d1:float | int, len1:float | int, len2:float | int, ro:float, revers=False) -> float | int:
    """
    Вычисляет массу болвана при заливке в цб кокиль.\n
    d1 - Диаметр кокиля.\n
    len1 - Длина кокиля.\n
    len2 - Длина отливки.\n
    ro - Плотность металла.\n
    reverse - Направление заливки [от большего диаметра к меньшему (False) default | обратное - как в стакан (True)].
    """

    diam_max_kok= len1/30+d1 # Диаметр у заливочной крышки
    if revers:
        diam_max= len2/30+d1 # Диаметр у заливочной крышки
        diam_avg = (d1 + diam_max)/2
    else:
        diam_min = (len1-len2)/30+d1
        diam_avg = (diam_max_kok + diam_min)/2
    return (diam_avg*0.985)**2*3.1415926*(len2*0.985)*ro/4_000_000
    
def get_diam2_kok(d1:int, len:int) -> tuple:
    """
    Принимает на вход диаметр кокиля и его длину.\n
    Возвращает: кортэж с всей инфой.
    """
    file = '//ad/технологи/Заказы/ЦБ ОСНАСТКА+ИЗЛОЖНИЦЫ.xlsx'
    book = op.load_workbook(file)
    ws = book['Изложницы']

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=9, max_row=45, values_only=True):
        if d1 == row[0] and len == row[2]:
            return row


pprint(round_val(105.5))